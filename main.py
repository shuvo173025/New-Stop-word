import openpyxl
import csv


my_test_word = open('test.csv', encoding="utf8")
my_test_word_reader = csv.reader(my_test_word)


word_list = []
rural_stop_word_list = []
temp = []


for i in my_test_word_reader:
    word_list.append(i)


wb = openpyxl.load_workbook("stop_word.xlsx")
ws = wb['Sheet1']
colomn = ws["A"]
for i in range(1,len(colomn)+1):
    C = 'A' + str(i)
    D = 'B' + str(i)
    x = ws[C].value
    for j in range(len(word_list)):
        if x == word_list[j][0]:
            y = word_list[j][1]
            ws[D].value = y

wb.save("stop_word.xlsx")

for i in range(1,len(colomn)+1):
    C = 'B' + str(i)
    D = 'A' + str(i)
    x = ws[C].value
    y= ws[D].value
    if x == None:
        rural_stop_word_list.append(y)
    else:
        rural_stop_word_list.append(x)


for i in rural_stop_word_list:
    z = '/'
    if z in i:
       rural_stop_word_list.remove(i)
       word = str(i).split('/')
       for w in word:
           rural_stop_word_list.append(w)
