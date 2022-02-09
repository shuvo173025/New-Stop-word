import openpyxl
import csv


my_test_word = open('test.csv', encoding="utf8")
my_test_word_reader = csv.reader(my_test_word)

stop_word = open('stop_word.csv', encoding="utf8")
stop_word_reader = csv.reader(stop_word)

word_list = []
new_word_list = []
stop_word_list = []
rural_stop_word_list = []
temp = []


for i in my_test_word_reader:
    word_list.append(i)


wb = openpyxl.load_workbook("stop_word.xlsx")
ws = wb['Form Responses 1']
colomn = ws["A"]
for i in range(1,len(colomn)+1):
    C = 'A' + str(i)
    D = 'B' + str(i)
    x = ws[C].value
    for j in range(len(word_list)):
        if x == word_list[j][0]:
            y = word_list[j][1]
            ws[D].value = y

wb.save("stop_word.xlsx")

for i in range(1,len(colomn)+1):
    C = 'B' + str(i)
    D = 'A' + str(i)
    x = ws[C].value
    y= ws[D].value
    if x == None:
        rural_stop_word_list.append(y)
    else:
        rural_stop_word_list.append(x)


for i in rural_stop_word_list:
    z = '/'
    if z in i:
       rural_stop_word_list.remove(i)
       word = str(i).split('/')
       for w in word:
           rural_stop_word_list.append(w)


# with open('stop_word.csv', 'w', encoding='UTF8', newline='') as w:
#     writer = csv.writer(w)
#     for i in rural_stop_word_list:
#         row_data = [i]
#         writer.writerow(row_data)



for i in stop_word_reader:
    stop_word_list.append(i[0])


wb = openpyxl.load_workbook("NLP_Data_Set.xlsx")
ws = wb['Form Responses 1']
colomn = ws["B"]

for i in range(2,len(colomn)+1):
    data = ''
    temp = ''
    C = 'B' + str(i)
    D = 'H' + str(i)
    x = ws[C].value
    word_list = list(x.split(" "))
    for j in word_list:
        if j not in stop_word_list:
            new_word_list.append(j)

    temp = ' '.join(new_word_list)
    data = temp.replace('/',' ').replace('?',' ').replace('।',' ').replace(',',' ').replace('!',' ').replace('-',' ')
    ws[D].value = data
    new_word_list = []

wb.save("NLP_Data_Set.xlsx")